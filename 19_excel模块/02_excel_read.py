# _*_ coding : utf-8 _*_
# @Time : 2024/7/17 20:53
# @Author : chalet.you
# @File : 02_excel_read
# @Project : bigdata_python_course
from openpyxl import load_workbook

wb = load_workbook("E:/youxuan/bigdata/IdeaProject_Study/datas/user1.xlsx")
ws = wb.sheetnames

for sheet_name in ws:
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=4, max_row=10, max_col=3, values_only=True):
        print(','.join(str(cell) for cell in row))
